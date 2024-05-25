import pandas as pd
from openpyxl import load_workbook

# Load the Excel file
file_path = "attendance.xlsx"  # Update this to the correct file path
excel_file = pd.ExcelFile(file_path)

# Initialize an empty DataFrame for consolidation
consolidated_df = pd.DataFrame()

# Iterate through each sheet in the Excel file
for sheet_name in excel_file.sheet_names:
    # Read the sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Clean up column names (trim whitespace, make lowercase, etc.)
    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")

    # Rename 'attendance' column to the sheet name (date)
    if "attendance" in df.columns:
        df = df.rename(columns={"attendance": sheet_name})

        # Keep only the relevant columns
        df = df[["roll_no", sheet_name]]

        # If consolidated_df is empty, initialize it with the current DataFrame
        if consolidated_df.empty:
            consolidated_df = df
        else:
            # Merge the current DataFrame with the consolidated DataFrame
            consolidated_df = pd.merge(consolidated_df, df, on="roll_no", how="outer")

# Check if the 'Consolidated' sheet exists and delete it if necessary
book = load_workbook(file_path)
if "Consolidated" in book.sheetnames:
    std = book["Consolidated"]
    book.remove(std)
    book.save(file_path)

# Write the consolidated DataFrame to a new 'Consolidated' sheet
with pd.ExcelWriter(file_path, engine="openpyxl", mode="a") as writer:
    consolidated_df.to_excel(writer, sheet_name="Consolidated", index=False)

print("Consolidation complete. Check the 'Consolidated' sheet in the Excel file.")
